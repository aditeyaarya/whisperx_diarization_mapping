import re
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook, Workbook

def append_mapping_to_excel(xlsx_path: Path,
                            df_names_row: pd.DataFrame,  # one row
                            df_codes_row: pd.DataFrame,  # one row
                            spk_tags: List[str],
                            key_col: str = "ID"):
    """
    Upsert into two sheets ('Names', 'Codes') in-place using openpyxl.
    - Primary key is key_col (default 'ID') matched case-sensitively in storage (we standardize writing).
    - Adds missing SPEAKER_xx columns while preserving order.
    """
    def _norm(s: str) -> str:
        return " ".join(str(s).split()).strip().lower() if s is not None else ""

    SPEAKER_COL_RE = re.compile(r'^\s*speaker[\s_\-]*0*(\d+)\s*$', re.IGNORECASE)
    NAMES_ALIASES = {"names", "name"}
    CODES_ALIASES = {"codes", "code"}

    def _pick_sheet(wb: Workbook, desired: str, aliases: set) -> str:
        norm_map = {_norm(n): n for n in wb.sheetnames}
        if _norm(desired) in norm_map:
            return norm_map[_norm(desired)]
        for alias in aliases:
            if alias in norm_map:
                return norm_map[alias]
        for n in wb.sheetnames:
            nn = _norm(n)
            if any(nn.startswith(a) for a in aliases) or nn.startswith(_norm(desired)):
                return n
        ws = wb.create_sheet(title=desired)
        return ws.title

    def _get_header(ws) -> List[str]:
        if ws.max_row >= 1:
            return [ws.cell(1, j).value for j in range(1, ws.max_column + 1)]
        return []

    def _ensure_key_and_speakers(ws, existing_cols: List[str], needed_spk_tags: List[str]) -> List[str]:
        cols = list(existing_cols) if existing_cols else []
        has_id = any(_norm(c) == _norm(key_col) for c in cols if c)
        if not has_id:
            cols = [key_col] + cols

        present_idxs = set()
        for c in cols:
            if not c: continue
            m = SPEAKER_COL_RE.match(str(c))
            if m: present_idxs.add(int(m.group(1)))

        for tag in needed_spk_tags:
            idx = int(str(tag).split("_")[1])
            if idx not in present_idxs:
                cols.append(f"SPEAKER_{idx:02d}")
        return cols

    def _index_map_by_header(ws, headers: List[str]) -> Dict[str, int]:
        current = _get_header(ws)
        if not current:
            for j, name in enumerate(headers, start=1):
                ws.cell(1, j, name)
            current = headers[:]
        else:
            have_norm = {_norm(c): True for c in current if c}
            for h in headers:
                if _norm(h) not in have_norm:
                    ws.cell(1, len(current) + 1, h)
                    current.append(h)
        current = _get_header(ws)
        return {str(c): i + 1 for i, c in enumerate(current)}

    def _find_key_col_idx(col_idx_map: Dict[str, int]) -> int:
        for name, j in col_idx_map.items():
            if _norm(name) == _norm(key_col):
                return j
        return 1

    def _find_row(ws, key_val: str, key_idx: int) -> Optional[int]:
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, key_idx).value == key_val:
                return r
        return None

    def _write_row(ws, col_idx_map: Dict[str, int], row_dict_std: Dict[str, Any]):
        key_idx = _find_key_col_idx(col_idx_map)
        key_val = str(row_dict_std.get(key_col, ""))
        r = _find_row(ws, key_val, key_idx)
        if r is None:
            r = max(ws.max_row + 1, 2)

        for hdr, j in col_idx_map.items():
            if _norm(hdr) == _norm(key_col):
                ws.cell(r, j, row_dict_std.get(key_col, ""))
                continue
            m = SPEAKER_COL_RE.match(str(hdr))
            if m:
                idx = int(m.group(1))
                std_tag = f"SPEAKER_{idx:02d}"
                ws.cell(r, j, row_dict_std.get(std_tag, ws.cell(r, j).value))
                continue
            # leave non-speaker columns untouched

    # open/create wb
    xlsx_path = Path(xlsx_path)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    if xlsx_path.exists():
        try:
            wb = load_workbook(xlsx_path)
        except Exception:
            wb = Workbook()
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            del wb["Sheet"]

    # sheets
    names_sheet_name = _pick_sheet(wb, "Names", NAMES_ALIASES)
    codes_sheet_name = _pick_sheet(wb, "Codes", CODES_ALIASES)
    ws_names = wb[names_sheet_name]
    ws_codes = wb[codes_sheet_name]

    # headers
    existing_names_cols = _get_header(ws_names)
    existing_codes_cols = _get_header(ws_codes)
    names_headers = _ensure_key_and_speakers(ws_names, existing_names_cols, spk_tags)
    codes_headers = _ensure_key_and_speakers(ws_codes, existing_codes_cols, spk_tags)
    col_idx_names = _index_map_by_header(ws_names, names_headers)
    col_idx_codes = _index_map_by_header(ws_codes, codes_headers)

    # rows as standardized dicts
    row_names_std = df_names_row.iloc[0].to_dict()
    row_codes_std = df_codes_row.iloc[0].to_dict()

    # write
    _write_row(ws_names, col_idx_names, row_names_std)
    _write_row(ws_codes, col_idx_codes, row_codes_std)

    wb.save(xlsx_path)
